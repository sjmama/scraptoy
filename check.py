import urllib.request
from bs4 import BeautifulSoup
import urllib.parse
import cv2 as cv
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def get_gray_shade(value):
    # 5단계 회색 음영으로 분류
    if value < 51:
        return 'FF333333'  # 매우 어두운 회색
    elif value < 102:
        return 'FF666666'  # 어두운 회색
    elif value < 153:
        return 'FF999999'  # 중간 회색
    elif value < 204:
        return 'FFCCCCCC'  # 밝은 회색
    else:
        return 'FFFFFFFF'  # 매우 밝은 회색
    
ASCII_CHARS = np.array(["$@B%8&WM#*oahkbdpqwmZO0QLCJUYXzcvunxrjft/\|()1{}[]?-_+~<>i!lI;:,\"^`'. "])
np.set_printoptions(threshold=np.inf, linewidth=np.inf)
def num2char(x):
    return ASCII_CHARS[0][x]
    
# 찾으려는 검색어를 입력받습니다.
search = input("찾으려는거 : ")
encoded_search = urllib.parse.quote(search)
url = f"https://www.youtube.com/results?search_query={encoded_search}"
# 요청을 보내고 응답을 받습니다.
req = urllib.request.Request(url)
fp = urllib.request.urlopen(req)
source = fp.read()
fp.close()

# BeautifulSoup을 사용하여 HTML을 파싱합니다.
soup = BeautifulSoup(source, 'html.parser')
ssoup=str(soup)

video_div = ssoup.find("videoRenderer")
id=ssoup[video_div:video_div+50].split('"')[4]

url=f"https://img.youtube.com/vi/{id}/maxresdefault.jpg"
img=urllib.request.urlopen(url).read()

img = np.asarray(bytearray(img), dtype=np.uint8)

# numpy 배열을 OpenCV 이미지로 변환합니다.
img = cv.imdecode(img, cv.IMREAD_GRAYSCALE)

dst2 = cv.resize(img, dsize=(0, 0), fx=0.21, fy=0.5, interpolation=cv.INTER_LINEAR)
# cv.imshow('Image', dst2)
# cv.waitKey(0)
# cv.destroyAllWindows()
numimg=dst2[:]//51
ascii_img = np.vectorize(num2char)(numimg)
# with open("result.jpg", "wb") as f:
#     f.write(img)
# with open("result.txt", "wb") as f:
#     f.write(ascii_img)
    
wb = Workbook()
ws = wb.active

# 각 픽셀 값을 5단계 회색 음영으로 분류하여 엑셀 셀을 채웁니다.
for i in range(dst2.shape[0]):
    for j in range(dst2.shape[1]):
        gray_value = dst2[i, j]
        fill_color = get_gray_shade(gray_value)
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        cell = ws.cell(row=i+1, column=j+1)
        cell.fill = fill

# 엑셀 파일 저장
wb.save(f"{search}.xlsx")